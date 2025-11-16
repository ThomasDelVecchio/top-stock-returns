#!/usr/bin/env python3

import pandas as pd
import yfinance as yf
from datetime import timedelta
from math import sqrt
from openpyxl.styles import Alignment  # for Excel note formatting

# ==========================================================
# CONFIG
# ==========================================================
STOCKANALYSIS_URL = "https://stockanalysis.com/list/biggest-companies/"
TOP_N = 1000                      # target number of tickers

YFINANCE_PERIOD = "1y"
YFINANCE_BATCH_SIZE = 200         # tickers per yfinance request

HORIZONS_DAYS = {
    "1w": 7,
    "1m": 30,
    "3m": 90,
    "6m": 180,
}

MASTER_CSV = "top1000_returns_master.csv"
EXCEL_REPORT = "top1000_return_report.xlsx"


# ==========================================================
# SYMBOL NORMALIZATION FOR YAHOO
# ==========================================================
def to_yahoo_symbol(sym: str) -> str:
    """
    Convert StockAnalysis / 'dot' style tickers to Yahoo style.
    Examples: BRK.B -> BRK-B, PBR.A -> PBR-A, HEI.A -> HEI-A
    """
    sym = sym.strip().upper()

    overrides = {
        "BRK.B": "BRK-B",
        "BRK.A": "BRK-A",
        "BF.B": "BF-B",
        "HEI.A": "HEI-A",
        "PBR.A": "PBR-A",
    }
    if sym in overrides:
        return overrides[sym]

    # Generic rule: replace '.' with '-'
    return sym.replace(".", "-")


# ==========================================================
# 1) SCRAPE ~TOP 1,000 TICKERS FROM STOCKANALYSIS (MULTI-PAGE)
# ==========================================================
def fetch_top_tickers_from_stockanalysis(top_n: int = TOP_N, max_pages: int = 5):
    """
    Scrape StockAnalysis 'Biggest Companies' list.
    Walks pages ?page=1,2,... until we have at least top_n tickers
    or we run out of pages.
    """
    all_syms = []
    page = 1

    while len(all_syms) < top_n and page <= max_pages:
        url = STOCKANALYSIS_URL if page == 1 else f"{STOCKANALYSIS_URL}?page={page}"
        print(f"[STEP] Fetching table from StockAnalysis page {page}: {url}")

        try:
            tables = pd.read_html(url)
        except Exception as e:
            print(f"[WARN] Failed to read HTML table on page {page}: {e}")
            break

        if not tables:
            print(f"[WARN] No tables found on page {page}.")
            break

        df = tables[0]

        # Find symbol column
        symbol_col = None
        for c in df.columns:
            name = str(c).lower()
            if "symbol" in name or "ticker" in name:
                symbol_col = c
                break
        if symbol_col is None:
            if len(df.columns) >= 2:
                symbol_col = df.columns[1]
                print(f"[WARN] Symbol column not labelled clearly. Using '{symbol_col}' on page {page}.")
            else:
                raise ValueError(f"Could not identify symbol column in scraped table on page {page}.")

        syms = (
            df[symbol_col]
            .dropna()
            .astype(str)
            .str.strip()
            .str.upper()
            .tolist()
        )

        all_syms.extend(syms)
        print(f"[INFO] Collected {len(all_syms)} raw symbols after page {page}.")

        page += 1

    # De-duplicate while preserving order
    seen = set()
    clean = []
    for s in all_syms:
        if s and s not in seen:
            seen.add(s)
            clean.append(s)

    tickers = clean[:top_n]
    print(f"[INFO] Got {len(tickers)} unique tickers from StockAnalysis (target {top_n}).")
    print(f"[INFO] First 10 tickers: {tickers[:10]}")
    return tickers


# ==========================================================
# 2) DOWNLOAD PRICES FROM YAHOO (USING NORMALIZED SYMBOLS)
# ==========================================================
def download_price_history_batched(tickers, period=YFINANCE_PERIOD, batch_size=YFINANCE_BATCH_SIZE):
    """
    tickers: list of original symbols (e.g. BRK.B)
    We normalize to Yahoo symbols internally (e.g. BRK-B)
    but keep columns / index as the original tickers.
    More robust with extra debug prints.
    """
    print(f"[STEP] Downloading daily prices for {len(tickers)} tickers from Yahoo Finance...")
    all_closes = {}

    for i in range(0, len(tickers), batch_size):
        batch_orig = tickers[i:i + batch_size]
        batch_yahoo = [to_yahoo_symbol(t) for t in batch_orig]

        print(f"[INFO]  Batch {i // batch_size + 1}: {len(batch_orig)} tickers "
              f"({', '.join(batch_orig[:5])}...)")
        print(f"[DEBUG] Yahoo symbols sample: {', '.join(batch_yahoo[:5])}")

        try:
            data = yf.download(
                tickers=batch_yahoo,
                period=period,
                interval="1d",
                auto_adjust=True,
                group_by="ticker",
                progress=False,
                threads=True,
            )
        except Exception as e:
            print(f"[WARN] yf.download failed for batch {i // batch_size + 1}: {e}")
            continue

        if data is None or data.empty:
            print(f"[WARN] Empty price data returned for batch {i // batch_size + 1}. Skipping.")
            continue

        # Multi-ticker case
        if isinstance(data.columns, pd.MultiIndex):
            for orig, ysym in zip(batch_orig, batch_yahoo):
                try:
                    if (ysym, "Close") in data.columns:
                        series = data[(ysym, "Close")].dropna()
                    elif (ysym, "Adj Close") in data.columns:
                        series = data[(ysym, "Adj Close")].dropna()
                    else:
                        continue
                    if not series.empty:
                        all_closes[orig] = series
                    else:
                        print(f"[DEBUG] No non-NA prices for {orig} ({ysym}) in this batch.")
                except KeyError:
                    print(f"[DEBUG] {ysym} not found in returned columns for this batch.")
                    continue
        else:
            # Single-ticker case (should be rare here but keep it)
            orig = batch_orig[0]
            if "Close" in data.columns:
                series = data["Close"].dropna()
            elif "Adj Close" in data.columns:
                series = data["Adj Close"].dropna()
            else:
                series = None

            if series is not None and not series.empty:
                all_closes[orig] = series
            else:
                print(f"[DEBUG] No non-NA prices for single-ticker batch {orig}.")

    if not all_closes:
        raise RuntimeError(
            "No price data downloaded at all. Likely causes:\n"
            "  - No internet connection\n"
            "  - VPN/firewall blocking Yahoo Finance\n"
            "  - Temporary Yahoo issue (try again in a bit)\n"
        )

    prices = pd.DataFrame(all_closes).sort_index()
    print(f"[INFO] Price history shape: {prices.shape} (rows = days, cols = tickers)")
    return prices


# ==========================================================
# 3) COMPUTE HORIZON RETURNS
# ==========================================================
def compute_horizon_returns(prices, horizons_days):
    """
    Returns DataFrame:
      index  = ticker (original symbols)
      cols   = '1w', '1m', '3m', '6m'
      values = % return (e.g. 12.34 means +12.34%)
    """
    print("[STEP] Computing horizon returns...")
    prices = prices.sort_index()
    last_date = prices.index.max()
    last_prices = prices.loc[last_date]

    result = {}

    for label, days in horizons_days.items():
        anchor_date = last_date - timedelta(days=days)
        eligible_dates = prices.index[prices.index <= anchor_date]

        if len(eligible_dates) == 0:
            print(f"[WARN] Not enough history for {label} ({days} days). Filling NaNs.")
            result[label] = pd.Series(index=prices.columns, dtype=float)
            continue

        anchor_prices = prices.loc[eligible_dates.max()]
        horizon_ret = (last_prices / anchor_prices - 1.0) * 100.0
        result[label] = horizon_ret

    returns_df = pd.DataFrame(result)
    returns_df.index.name = "ticker"
    return returns_df


# ==========================================================
# 3b) TREND SCORES (ALL TICKERS)
# ==========================================================
def compute_trend_scores(returns_df):
    """
    Compute per-horizon percentile scores and composite_score for ALL tickers.

    Returns DataFrame with columns:
      1w_score, 1m_score, 3m_score, 6m_score, composite_score
    """
    if returns_df.empty:
        return pd.DataFrame()

    df = returns_df.copy()
    n = len(df)
    scores = pd.DataFrame(index=df.index)

    if n > 1:
        for col in df.columns:
            ranks = df[col].rank(ascending=False, method="average")
            scores[f"{col}_score"] = 1.0 - (ranks - 1.0) / (n - 1.0)
    else:
        for col in df.columns:
            scores[f"{col}_score"] = 1.0

    score_cols = [c for c in scores.columns if c.endswith("_score")]
    scores["composite_score"] = scores[score_cols].mean(axis=1)

    return scores


# ==========================================================
# 3c) VOLATILITY & RISK-ADJUSTED SCORE
# ==========================================================
def compute_vol_and_risk_adjusted(prices, returns_df):
    """
    Compute simple daily-vol-based risk adjustment:

      vol_6m = std dev of daily returns over last ~6m
      risk_adj_score = (6m_return% / 100) / vol_6m

    Returns DataFrame with vol_6m and risk_adj_score.
    """
    if prices.empty or returns_df.empty:
        return pd.DataFrame(index=returns_df.index)

    prices = prices.sort_index()
    last_date = prices.index.max()
    anchor_date = last_date - timedelta(days=HORIZONS_DAYS["6m"])
    price_6m = prices[prices.index >= anchor_date]

    daily_rets = price_6m.pct_change().dropna()
    vol_6m = daily_rets.std()  # per ticker, fraction (e.g. 0.02 = 2% daily)

    out = pd.DataFrame(index=returns_df.index)
    out["vol_6m"] = vol_6m.reindex(returns_df.index)

    # convert 6m % return to fraction
    sixm_frac = returns_df["6m"] / 100.0
    out["risk_adj_score"] = sixm_frac / out["vol_6m"]
    return out


# ==========================================================
# 3d) MOVING AVERAGE FLAGS (50d / 200d)
# ==========================================================
def compute_ma_flags(prices):
    """
    Compute 50d and 200d simple moving averages and flags:
      above_50d, above_200d (True/False)
    """
    if prices.empty:
        return pd.DataFrame()

    prices = prices.sort_index()
    ma50 = prices.rolling(window=50).mean()
    ma200 = prices.rolling(window=200).mean()

    last_price = prices.iloc[-1]
    last_ma50 = ma50.iloc[-1]
    last_ma200 = ma200.iloc[-1]

    out = pd.DataFrame(index=prices.columns)
    out.index.name = "ticker"
    out["above_50d"] = last_price > last_ma50
    out["above_200d"] = last_price > last_ma200

    return out


# ==========================================================
# 3e) PULLBACK & NEW-MOMENTUM FLAGS
# ==========================================================
def compute_pullback_flag(returns_df,
                          min_3m=25.0,
                          min_6m=40.0):
    """
    Pullback idea:
      - Strong 3m and 6m trend
      - Recent 1w or 1m <= 0 (dip in a strong uptrend)
    """
    df = returns_df.copy()
    strong = (df["3m"] >= min_3m) & (df["6m"] >= min_6m)
    recent_dip = (df["1w"] <= 0.0) | (df["1m"] <= 0.0)
    flag = strong & recent_dip
    return flag.rename("pullback_flag")


def compute_new_momentum_flag(returns_df):
    """
    New momentum idea:
      - shorter horizons stronger than longer ones, roughly:
          1m >= 3m/3
          3m >= 6m/3
      - and all positive
    This is intentionally loose: you're just looking for acceleration.
    """
    df = returns_df.copy()
    all_pos = (df["1w"] > 0) & (df["1m"] > 0) & (df["3m"] > 0) & (df["6m"] > 0)

    cond_1 = df["1m"] >= df["3m"] / 3.0
    cond_2 = df["3m"] >= df["6m"] / 3.0

    flag = all_pos & cond_1 & cond_2
    return flag.rename("new_momentum_flag")


# ==========================================================
# 3f) SECTOR LOOKUP
# ==========================================================
def fetch_sectors_for_tickers(tickers):
    """
    Best effort sector lookup via yfinance.Ticker(info).
    """
    sectors = {}
    print("[STEP] Fetching sectors from Yahoo (best effort)...")
    for t in tickers:
        ysym = to_yahoo_symbol(t)
        try:
            info = yf.Ticker(ysym).info
            sector = info.get("sector", "Unknown")
        except Exception:
            sector = "Unknown"
        sectors[t] = sector
    return pd.Series(sectors, name="sector")


# ==========================================================
# 3g) FINAL SCORE (COMBINED VIEW)
# ==========================================================
def compute_final_scores(analytics_df):
    """
    Combine multiple signals into a single final_score:

      Inputs (per ticker):
        composite_score  (0–1)
        risk_adj_score   (numeric, risk-adjusted 6m return)
        new_momentum_flag (bool)
        pullback_flag     (bool)

      Steps:
        - Convert risk_adj_score into a 0–1 percentile (risk_adj_pct)
        - Flags -> 0/1
        - final_score = 0.4 * composite_score
                        + 0.3 * risk_adj_pct
                        + 0.2 * new_mom_flag
                        + 0.1 * pullback_flag
    """
    df = analytics_df.copy()
    n = len(df)
    if n == 0:
        df["final_score"] = pd.NA
        return df

    # Risk-adjusted percentile (higher is better)
    if "risk_adj_score" in df.columns and df["risk_adj_score"].notna().any():
        ranks = df["risk_adj_score"].rank(ascending=False, method="average")
        if n > 1:
            df["risk_adj_pct"] = 1.0 - (ranks - 1.0) / (n - 1.0)
        else:
            df["risk_adj_pct"] = 1.0
    else:
        df["risk_adj_pct"] = 0.0

    # Flags to 0/1
    for col in ["new_momentum_flag", "pullback_flag"]:
        if col in df.columns:
            df[col] = df[col].fillna(False).astype(int)
        else:
            df[col] = 0

    # Composite is already 0–1
    df["composite_score"] = df["composite_score"].fillna(0.0)

    df["final_score"] = (
        0.4 * df["composite_score"]
        + 0.3 * df["risk_adj_pct"]
        + 0.2 * df["new_momentum_flag"]
        + 0.1 * df["pullback_flag"]
    )

    return df


# ==========================================================
# 4) RANKED VIEWS + SAVING (HUMAN-FRIENDLY SHEETS)
# ==========================================================
def build_ranked_views(returns_df):
    ranked_views = {}
    for col in returns_df.columns:
        ranked_views[col] = returns_df.sort_values(col, ascending=False)
    return ranked_views


def _yes_no_map(val):
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if isinstance(val, (int, float)):
        return "Yes" if val >= 0.5 else "No"
    return "No"


def save_report(
    returns_df,
    ranked_views,
    top_trend_df,
    top_final_df,
    top_sector_df,
    master_csv=MASTER_CSV,
    excel_path=EXCEL_REPORT,
):
    print("[STEP] Saving master CSV and Excel report...")

    # ---------- CSV (keep values as % numbers, not decimal) ----------
    returns_df_csv = returns_df.rename(
        columns={
            "1w": "1w_return_pct",
            "1m": "1m_return_pct",
            "3m": "3m_return_pct",
            "6m": "6m_return_pct",
        }
    )
    returns_df_csv.to_csv(master_csv, float_format="%.4f")
    print(f"[OUT] Master CSV saved to {master_csv}")

    # ---------- Excel ----------
    # For Excel, convert returns to decimal (0.1234) for % formatting
    def display_returns(df):
        df_disp = df.copy() / 100.0
        df_disp.columns = ["1W %", "1M %", "3M %", "6M %"]
        return df_disp

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # 1) Master sheet
        display_returns(returns_df).to_excel(writer, sheet_name="Master")

        # 2) Horizon sheets
        for horizon, df_view in ranked_views.items():
            sheet_name = f"Top_{horizon}"
            display_returns(df_view).to_excel(writer, sheet_name=sheet_name)

        # 3) Top_Trend_10 (enriched, but renamed & cleaned)
        if top_trend_df is not None and not top_trend_df.empty:
            trend_disp = top_trend_df.copy()

            # Convert returns to decimals
            for c in ["1w", "1m", "3m", "6m"]:
                if c in trend_disp.columns:
                    trend_disp[c] = trend_disp[c] / 100.0

            # Convert flags to Yes/No
            for col in ["new_momentum_flag", "pullback_flag", "above_50d", "above_200d"]:
                if col in trend_disp.columns:
                    trend_disp[col] = trend_disp[col].map(_yes_no_map)

            # Rename columns to be readable
            trend_rename = {
                "composite_score": "Trend_Strength_Score",
                "risk_adj_score": "Risk_Adjusted_6M",
                "risk_adj_pct": "Risk_Adjusted_Percentile",
                "new_momentum_flag": "Is_Trend_Accelerating",
                "pullback_flag": "Is_Strong_Trend_Pulled_Back",
                "above_50d": "Above_50_Day_MA",
                "above_200d": "Above_200_Day_MA",
                "sector": "Sector",
                "1w": "1W %",
                "1m": "1M %",
                "3m": "3M %",
                "6m": "6M %",
                "final_score": "Final_Composite_Score",
            }
            trend_disp = trend_disp.rename(columns=trend_rename)

            trend_disp.to_excel(writer, sheet_name="Top_Trend_10", startrow=1)

        # 4) Top_Final_10 (full info, readable names)
        lite_disp = None
        if top_final_df is not None and not top_final_df.empty:
            final_disp = top_final_df.copy()

            for c in ["1w", "1m", "3m", "6m"]:
                if c in final_disp.columns:
                    final_disp[c] = final_disp[c] / 100.0

            for col in ["new_momentum_flag", "pullback_flag", "above_50d", "above_200d"]:
                if col in final_disp.columns:
                    final_disp[col] = final_disp[col].map(_yes_no_map)

            final_rename = {
                "composite_score": "Trend_Strength_Score",
                "risk_adj_score": "Risk_Adjusted_6M",
                "risk_adj_pct": "Risk_Adjusted_Percentile",
                "new_momentum_flag": "Is_Trend_Accelerating",
                "pullback_flag": "Is_Strong_Trend_Pulled_Back",
                "above_50d": "Above_50_Day_MA",
                "above_200d": "Above_200_Day_MA",
                "sector": "Sector",
                "1w": "1W %",
                "1m": "1M %",
                "3m": "3M %",
                "6m": "6M %",
                "final_score": "Final_Composite_Score",
            }
            final_disp = final_disp.rename(columns=final_rename)

            final_disp.to_excel(writer, sheet_name="Top_Final_10", startrow=1)

            # Build a Lite view for non-experts
            wanted_cols = [
                "Sector",
                "Final_Composite_Score",
                "Trend_Strength_Score",
                "1W %",
                "1M %",
                "3M %",
                "6M %",
                "Is_Trend_Accelerating",
                "Is_Strong_Trend_Pulled_Back",
                "Above_200_Day_MA",
            ]
            lite_cols = [c for c in wanted_cols if c in final_disp.columns]
            lite_disp = final_disp[lite_cols]
            lite_disp.to_excel(writer, sheet_name="Top_Final_10_Lite", startrow=1)

        # 5) Top_Sector_Leaders (simplified but still more advanced)
        if top_sector_df is not None and not top_sector_df.empty:
            sector_disp = top_sector_df.copy()

            for c in ["1w", "1m", "3m", "6m"]:
                if c in sector_disp.columns:
                    sector_disp[c] = sector_disp[c] / 100.0

            for col in ["new_momentum_flag", "pullback_flag", "above_50d", "above_200d"]:
                if col in sector_disp.columns:
                    sector_disp[col] = sector_disp[col].map(_yes_no_map)

            sector_rename = {
                "composite_score": "Trend_Strength_Score",
                "risk_adj_score": "Risk_Adjusted_6M",
                "risk_adj_pct": "Risk_Adjusted_Percentile",
                "new_momentum_flag": "Is_Trend_Accelerating",
                "pullback_flag": "Is_Strong_Trend_Pulled_Back",
                "above_50d": "Above_50_Day_MA",
                "above_200d": "Above_200_Day_MA",
                "sector": "Sector",
                "1w": "1W %",
                "1m": "1M %",
                "3m": "3M %",
                "6m": "6M %",
                "final_score": "Final_Composite_Score",
            }
            sector_disp = sector_disp.rename(columns=sector_rename)

            sector_disp.to_excel(writer, sheet_name="Top_Sector_Leaders", startrow=1)

        # 6) ReadMe sheet explaining each tab in human terms
        readme_sheets = [
            "Master",
            "Top_1w",
            "Top_1m",
            "Top_3m",
            "Top_6m",
            "Top_Trend_10",
            "Top_Final_10_Lite",
            "Top_Final_10",
            "Top_Sector_Leaders",
        ]
        readme_desc = [
            "All stocks with 1-week, 1-month, 3-month, and 6-month returns. This is the base data.",
            "Same as Master, sorted by 1-week return (best recent week at the top).",
            "Same as Master, sorted by 1-month return.",
            "Same as Master, sorted by 3-month return.",
            "Same as Master, sorted by 6-month return.",
            "Top 10 multi-horizon trend leaders with detailed scores, risk, sector, and trend health.",
            "Simple view for non-experts: top 10 overall ideas with sector, scores, returns, and easy Yes/No flags.",
            "Full detail view of the final top 10 ideas with all advanced columns.",
            "Top 3 leaders in each sector based on trend strength, to see where leadership is concentrated.",
        ]
        readme_data = {"Sheet": readme_sheets, "Description": readme_desc}
        pd.DataFrame(readme_data).to_excel(writer, sheet_name="ReadMe", index=False)

        wb = writer.book

        # ----- Add explanatory notes -----
        if "Top_Trend_10" in wb.sheetnames:
            ws = wb["Top_Trend_10"]
            note_text = (
                "Note: Trend_Strength_Score shows how strong this stock's performance has been across 1W, 1M, 3M, and 6M "
                "compared to the full list (closer to 100% = stronger). 'Is_Trend_Accelerating' and 'Is_Strong_Trend_Pulled_Back' "
                "are simple Yes/No flags to indicate whether momentum is picking up or a strong trend has recently dipped."
            )
            ws["A1"] = note_text
            max_merge_col = min(10, ws.max_column if ws.max_column else 10)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_merge_col)
            ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")

        if "Top_Final_10_Lite" in wb.sheetnames:
            ws = wb["Top_Final_10_Lite"]
            note_text = (
                "Simple top-10 list for non-experts. Focus on: Final_Composite_Score (higher = more attractive), "
                "the 1W/1M/3M/6M returns, and the Yes/No flags. "
                "'Is_Trend_Accelerating' = momentum picking up. "
                "'Is_Strong_Trend_Pulled_Back' = strong multi-month trend that has recently dipped."
            )
            ws["A1"] = note_text
            max_merge_col = min(10, ws.max_column if ws.max_column else 10)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_merge_col)
            ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")

        if "Top_Sector_Leaders" in wb.sheetnames:
            ws = wb["Top_Sector_Leaders"]
            note_text = (
                "Shows the strongest names in each sector based on trend strength. "
                "Use this to see which sectors are leading the market and to avoid over-concentrating in a single theme."
            )
            ws["A1"] = note_text
            max_merge_col = min(8, ws.max_column if ws.max_column else 8)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_merge_col)
            ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")

        # ----- Apply % formatting ONLY to obvious percentage columns -----
        percent_like_keywords = ["%", "Score", "Percentile"]
        for ws in wb.worksheets:
            if ws.title in ("ReadMe",):
                continue
            headers = {cell.column: cell.value for cell in ws[2]}  # row 2 is header row
            for row in ws.iter_rows(min_row=3, min_col=2):
                for cell in row:
                    header = headers.get(cell.column, "")
                    if isinstance(cell.value, (int, float)) and any(
                        kw in str(header) for kw in percent_like_keywords
                    ):
                        cell.number_format = "0.00%"

    print(f"[OUT] Excel report saved to {excel_path}")


def print_top_snippets(returns_df, top_n=20):
    print("\n================ TOP MOVERS ================")
    for col in returns_df.columns:
        print(f"\n--- Top {top_n} by {col} return (percent points) ---")
        subset = returns_df.sort_values(col, ascending=False)
        print(subset.head(top_n).round(2))


# ==========================================================
# MAIN
# ==========================================================
def main():
    print("[START] Script started.")

    # 1) Scrape up to 1,000 tickers (multi-page)
    tickers = fetch_top_tickers_from_stockanalysis(TOP_N)
    print(f"[INFO] Using {len(tickers)} tickers.")

    # 2) Download price history
    prices = download_price_history_batched(tickers, period=YFINANCE_PERIOD)

    # 3) Compute returns
    returns_df = compute_horizon_returns(prices, HORIZONS_DAYS)
    returns_df = returns_df.dropna(how="all")

    # 4) Base ranked views
    ranked_views = build_ranked_views(returns_df)

    # 5) Trend scores (all tickers)
    trend_scores = compute_trend_scores(returns_df)

    # 6) Vol & risk-adjusted scores
    vol_risk = compute_vol_and_risk_adjusted(prices, returns_df)

    # 7) Moving-average flags
    ma_flags = compute_ma_flags(prices)

    # 8) Pullback & new-momentum flags
    pullback_flag = compute_pullback_flag(returns_df)
    new_mom_flag = compute_new_momentum_flag(returns_df)

    # 9) Sector (best-effort)
    sectors = fetch_sectors_for_tickers(returns_df.index)

    # 10) Assemble analytics master DataFrame
    analytics = trend_scores.join(vol_risk, how="left")
    analytics = analytics.join(pullback_flag, how="left")
    analytics = analytics.join(new_mom_flag, how="left")
    analytics = analytics.join(sectors, how="left")
    analytics = analytics.join(ma_flags, how="left")
    analytics = analytics.join(returns_df, how="left")

    # 11) Final combined score
    analytics = compute_final_scores(analytics)

    # Filter out junky names: at least 2 positive horizons
    positive_horizons = (returns_df > 0).sum(axis=1) >= 2

    # 12) Top trend (by composite_score)
    top_trend_df = (
        analytics[positive_horizons]
        .sort_values("composite_score", ascending=False)
        .head(10)
    )

    # 13) Top final recommendations (by final_score)
    top_final_df = (
        analytics[positive_horizons]
        .sort_values("final_score", ascending=False)
        .head(10)
    )

    # 14) Top sector leaders (top 3 per sector by composite_score)
    sector_filtered = analytics[positive_horizons & analytics["sector"].notna()]
    sector_filtered = sector_filtered[sector_filtered["sector"] != "Unknown"]
    top_sector_list = []
    for sector, group in sector_filtered.groupby("sector"):
        top3 = group.sort_values("composite_score", ascending=False).head(3)
        top_sector_list.append(top3)
    if top_sector_list:
        top_sector_df = pd.concat(top_sector_list)
    else:
        top_sector_df = pd.DataFrame()

    # 15) Console preview
    print_top_snippets(returns_df, top_n=20)
    if not top_final_df.empty:
        print("\n=========== TOP 10 FINAL RECOMMENDATIONS (final_score desc) ===========")
        cols_to_show = [
            "final_score",
            "composite_score",
            "risk_adj_score",
            "new_momentum_flag",
            "pullback_flag",
            "above_50d",
            "above_200d",
            "sector",
            "1w",
            "1m",
            "3m",
            "6m",
        ]
        cols_to_show = [c for c in cols_to_show if c in top_final_df.columns]
        print(top_final_df[cols_to_show].round(3))

    # 16) Save report
    save_report(
        returns_df,
        ranked_views,
        top_trend_df,
        top_final_df,
        top_sector_df,
        MASTER_CSV,
        EXCEL_REPORT,
    )

    print("[DONE] All finished.")


# ------------------------------------------------------------------
# 8) OPTIONAL: Copy outputs to Drive / local folders (Windows-safe)
# ------------------------------------------------------------------
try:
    import os, shutil, time

    def safe_copy(src, dst, max_retries=5, delay=1.0):
        """
        Copy file with retries to avoid Windows 'file in use' (WinError 32).
        """
        for attempt in range(1, max_retries + 1):
            try:
                shutil.copy2(src, dst)
                print(f"[COPY] {src} -> {dst}")
                return
            except PermissionError as e:
                msg = str(e)
                if "being used by another process" in msg or "WinError 32" in msg:
                    if attempt < max_retries:
                        print(f"[WARN] {src} locked (attempt {attempt}/{max_retries}). Retrying in {delay}s...")
                        time.sleep(delay)
                        continue
                    else:
                        print(f"[WARN] Gave up copying {src} after {max_retries} attempts.")
                        return
                else:
                    print(f"[WARN] Could not copy {src} -> {dst}: {e}")
                    return
            except Exception as e:
                print(f"[WARN] Could not copy {src} -> {dst}: {e}")
                return

    # Collect existing output files
    files_to_copy = []
    if os.path.exists(MASTER_CSV):
        files_to_copy.append(MASTER_CSV)
    if os.path.exists(EXCEL_REPORT):
        files_to_copy.append(EXCEL_REPORT)

    # A) Copy to your Windows project folder
    windows_project_folder = r"C:\Users\Tommy\top_stock_returns"
    if os.path.isdir(windows_project_folder):
        for fname in files_to_copy:
            dst = os.path.join(windows_project_folder, os.path.basename(fname))
            safe_copy(fname, dst)

    # B) Copy to Google Drive on Windows (G:)
    windows_drive_path = r"G:\My Drive\Top Stocks Output"
    if os.path.isdir(windows_drive_path):
        for fname in files_to_copy:
            dst = os.path.join(windows_drive_path, os.path.basename(fname))
            safe_copy(fname, dst)

    # C) Copy to Colab Google Drive folder (if running in Google Colab)
    colab_drive_path = "/content/drive/MyDrive/Top Stocks Output"
    if os.path.isdir(colab_drive_path):
        for fname in files_to_copy:
            dst = os.path.join(colab_drive_path, os.path.basename(fname))
            safe_copy(fname, dst)

except Exception as outer_err:
    print(f"[WARN] Output copy step failed: {outer_err}")


if __name__ == "__main__":
    import traceback
    try:
        main()
    except Exception as e:
        print("\n[ERROR] An exception occurred:")
        print(e)
        traceback.print_exc()
